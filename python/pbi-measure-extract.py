import json
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# Define the function to extract measures
def extract_measures(bim_file_path):
    # Load the BIM file content
    with open(bim_file_path, 'r', encoding='utf-8') as file:
        bim_content = json.load(file)

    # Extract measures
    measures = []
    for table in bim_content['model']['tables']:
        for measure in table.get('measures', []):
            measure_table = table.get('name', 'N/A')
            measure_displayFolder = measure.get('displayFolder', 'N/A')
            measure_name = measure.get('name', 'N/A')
            measure_expression = measure.get('expression', 'N/A')
            measure_format_string = measure.get('formatString', 'N/A')
            measure_description = measure.get('description', 'N/A')
            measures.append({
                'Table': measure_table,
                'Folder': measure_displayFolder,
                'Measure Name': measure_name,
                'Definition': measure_expression,
                'Description': measure_description,
                'Format String': measure_format_string
            })

    # Create a DataFrame
    return pd.DataFrame(measures)

def format_measure_expression(expression):
    # when stored as a list, convert to string with new lines
    if isinstance(expression, list):
        expression = '\n'.join(expression)
    return expression

# Initialize Tkinter
root = tk.Tk()
root.withdraw()  # Hide the main window

# Show an "Open" dialog box and return the path to the selected file
bim_file_path = filedialog.askopenfilename(
    title="Select BIM file",
    filetypes=(("BIM files", "*.bim"), ("All files", "*.*"))
)

if bim_file_path:  # Proceed only if a file was selected
    measures_df = extract_measures(bim_file_path)
    measures_df['Definition'] = measures_df['Definition'].apply(format_measure_expression)
    measures_df['Description'] = measures_df['Description'].apply(format_measure_expression)
  
    # Show a "Save" dialog box to specify the path for the output Excel file
    output_excel_path = filedialog.asksaveasfilename(
        title="Save Excel file",
        filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")),
        defaultextension=".xlsx"
    )
    
    # Check if a file path was provided
    if output_excel_path:
        with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
            measures_df.to_excel(writer, index=False, sheet_name='Measures')
            
            # Access the openpyxl workbook and worksheet objects
            workbook = writer.book
            worksheet = writer.sheets['Measures']

            # Enable text wrapping for the 'Description' and 'Definition' columns
            for column in measures_df.columns:
                if column in ['Description', 'Definition']:
                    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=measures_df.columns.get_loc(column) + 1, max_col=measures_df.columns.get_loc(column) + 1):
                        for cell in row:
                            cell.alignment = Alignment(wrapText=True)
            
            # Get the Excel range of the DataFrame
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            data_range = f"A1:{chr(64 + max_col)}{max_row}"
            
            # Create a table over the DataFrame range
            table = Table(displayName="MeasuresTable", ref=data_range)
            
            # Add a default table style with options
            style = TableStyleInfo(name="TableStyleLight16", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            
            # Add the table to the worksheet
            worksheet.add_table(table)

            # Auto-size column widths
            for column in measures_df.columns:
                if column in ['Description', 'Definition']:
                    # Set a hardcoded width for 'Description' and 'Definition' columns
                    hardcoded_width = 90  # Example width, adjust as needed
                    worksheet.column_dimensions[get_column_letter(measures_df.columns.get_loc(column) + 1)].width = hardcoded_width
                else:
                    # Auto-size for other columns
                    max_length = max(measures_df[column].astype(str).map(len).max(), len(column))
                    worksheet.column_dimensions[get_column_letter(measures_df.columns.get_loc(column) + 1)].width = max_length * 1.2

    print("Process completed successfully.")
